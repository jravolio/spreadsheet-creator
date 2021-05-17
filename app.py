from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active


#functions
class colors:
    cyan = '\033[1;36m'
    white = '\033[1;97m'
    yellow = '\033[1;33m'
    end = '\033[0m'
    green = '\033[92m'
    red = '\033[91m'

def criar_pagina_planilha():
    pagina = input('Digite o nome da página: ')
    wb.create_sheet(pagina)


def limpar_terminal():
    os.system('cls' if os.name =='nt' else clear)


limpar_terminal()

#Bem vindo
print(colors.cyan + 'Bem vindo!\nPara começar crie uma nova página dentro da planilha.' + colors.end)
criar_pagina_planilha()

while True:
    resposta_criar_planilha = input('Deseja criar mais uma página nesta planilha? (s/n): ')
    if resposta_criar_planilha.lower() == 's' or resposta_criar_planilha.lower() =='sim':
        criar_pagina_planilha()
    elif resposta_criar_planilha.lower() =='n' or resposta_criar_planilha.lower() =='não' or resposta_criar_planilha.lower() =='nao':
        del wb['Sheet']
        limpar_terminal()
        print(colors.yellow + str(wb.sheetnames) + colors.end)
        break
    else:
        print(colors.red + 'Por favor digite uma resposta válida!' + colors.end)


while True:
    while True:
        try:
            planilha_a_manipular = input('Escolha a planilha que deseja manipular: ')
            planilha_selecionada = wb[planilha_a_manipular]
            lista = ([])
            break
        except KeyError as planilha_inexistente:
            print(colors.red + 'Digite uma planilha existente!' + colors.end)

    while True:
        lista.append(input('Digite um nome para seu cabeçalho: '))
        resposta_cabecalho = input(colors.green + 'Deseja adicionar mais uma coluna? (s/n): '+ colors.end)
        if resposta_cabecalho.lower() == 's' or resposta_cabecalho.lower() =='sim':
            pass
        elif resposta_cabecalho.lower() =='n' or resposta_cabecalho.lower() =='nao' or resposta_cabecalho.lower() =='não':
            planilha_selecionada.append(lista)
            break
        else:
            print(colors.red + 'Digite uma resposta válida! Apagando dados de cabeçalho escritos posteriormente!' + colors.end)
            lista = ([])
        
    adicionar_mais_dados = input('Deseja adicionar dados a essa planilha? (s/n): ')
    limpar_terminal()
    if adicionar_mais_dados.lower() == 's':
        pass
    elif adicionar_mais_dados.lower() == 'n':
        break
    else:
        print(colors.red + 'Não foi possível entender sua resposta, Finalizando programa...' + colors.end)
        break

    print(colors.yellow + f'Você está manipulando a planilha: {planilha_selecionada}'+ colors.end)
    lista = ([])

    while True:
        dados_a_adicionar = input(colors.yellow + 'Digite os dados a serem adicionados a uma nova linha, separados por vírgula: ' + colors.end)
        lista.append(dados_a_adicionar.split(','))
        resposta_adicionar_nova_linha = input('Adicionar nova linha? (s/n): ')
        if resposta_adicionar_nova_linha.lower() == 's':
            pass
        elif resposta_adicionar_nova_linha.lower() =='n':
            for dados in lista:
                planilha_selecionada.append(dados)
            limpar_terminal()
            break
        
    print(colors.yellow + str(wb.sheetnames) + colors.end)
    manipular2 = input(colors.cyan + 'Deseja Manipular mais alguma planilha? (s/n): ' + colors.end)
    if manipular2.lower() == 's':
        pass
    elif manipular2.lower() =='n':
        limpar_terminal()
        break
    else:
        print(colors.red + 'Não foi possível identificar sua resposta, parando a execução do programa!' + colors.end)


nome_para_salvar = input('Qual o nome deseja salvar sua planilha? ')

#row - linha
# column - coluna
wb.save(f'{nome_para_salvar}.xlsx')

print(colors.green + 'Planilha salva com sucesso!' + colors.end)

